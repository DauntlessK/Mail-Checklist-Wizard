import sys
import openpyxl, os
import win32print
import pandas as pd
import numpy as np
import pathlib
import re
#import tabula as tb
from openpyxl import load_workbook, Workbook
import win32com.client as win32
##Authors: Anthony Mikinka & Kyle Breen-Bondie
##Revised by Michelle Ahmed 
##Version .41


############ Extracting Data from Job Ticket PDF ############ 

#job_tik_pdf = 'JobTicket145752.pdf'
#job_data = tb.read_pdf(job_tik_pdf, pages = '3')
#type(job_data) #returns list
#job_df = pd.DataFrame(np.concatenate(job_data)) #turns list into dataframe
#print(job_df)
#job_df.dropna(inplace=True) #drops NaNp values

############ Accessing Finished Data File ############

#Break Mark (0) #Full Name (1) #Business (2) #Address Line 1 (3) #Address Line 2 (4) #City State ZIP Code (5) #IM Barcode (6) 
#Numeric IM Barcode (7) #Endorsement Line (8) #Sort Position (9) #Bundle Number (10) #Tray Number (11) #Pallet Number (12)

#filename = '166266 TAC Momentum PC.xlsx' #file = that used for  data
MARKUP_PERCENT = 0.03
files_to_upload = True
upload_num = 0
ALLOWABLE_MAILING_CLASSES = [       #List of mailing classes that can be entered, otherwise you cannot progress
    'Presort Standard',
    'Presort Standard Stamp',
    'Presort First Class',
    'Presort First Class Stamp',
    'Full First Class Stamp',
    'International Stamp',
    'Non Profit Stamp',
    'Non Profit',
    'Peridocal',
    'DP',
    'Meter'
]
EXPECTED_UPLOAD_COLUMNS = [         #Column names (exactly) that python is expecting within the upload xlsx file. In that order.
    'Brea',
    'Full Name',
    'Business',
    'Address Line 1',
    'Address Line 2',
    'City State Zip Code',
    'IM Barcode',
    'Numeric IM barcode',
    'Endorsement Line',
    'Sort Position'
]
filename_array = []     #contains a list of all file names in the current upload
job_info_array = []     #contains a list of all job info in current upload (filename - extension)
permit_num_array = []   #contains a list of permit nums - needed? -> only 1 permit num?
total_qty_array = []    #contains a list of all quantities
postage_array = []      #contains a list of all postage costs
undeliverables = False
client = ""
print("------ Python Mail Checklist Wizard ------")
client = input("Input client name: ")

#loop for each file
while files_to_upload:
    filename = input("Input file data name: ")
    if "Undeliverables" in filename:
        undeliverables = True
    job_info =  filename
    filename = filename + ".xlsx"
    filename_array.append(filename)
    job_info_array.append(job_info)
    job_num = re.search(r'^(\d+)\s', filename)
    if job_num:
        job_num = job_num.group(1)
    else:
        print("Error getting job number from file name.")
    permit_num = input("Input Permit number: ")
    permit_num_array.append(permit_num)
    postage = input("Input postage cost: $")

    ############ Input and Verify Mailing Class ############
    mailing_class_is_invalid = True
    while mailing_class_is_invalid:
        mailing_class = ""
        mailing_class = input("Input the mailing class: ")
        mail_class_txt = 'Mail Class: '
        mail_class = mail_class_txt + mailing_class
        for x in ALLOWABLE_MAILING_CLASSES:
            if mailing_class == x:
                mailing_class_is_invalid = False
                break
        if mailing_class_is_invalid:
            print("Error- input mailing class does not match allowable mailing classes. You entered: " + mailing_class)
            print("Allowable mailing classes: ", end = "")
            print(ALLOWABLE_MAILING_CLASSES)

    try: 
        data = pd.read_excel(filename_array[upload_num]) #loading excel file into pandas
    except:
        print("ERROR - Likely the wizard was not able to find the file in the current folder.")
        sys.exit()
    #data.to_excel(filename_array[upload_num] + ' original.xlsx')    #save original file
    print(data.head())

    data.fillna("na")

    ############ Excel Work for finished data file ############

    #removed bundle and tray
    #data['Bundle Number'] = 'P' + data['Bundle Number'].astype(str) #selecting bundle num column, appending the string to every row
    #data['Tray Number'] = 'T' + data['Tray Number'].astype(str) #selecting tray num column, appending the string to every row

    #data.drop(data.columns[[1]],axis = 1)
    data.to_excel(filename_array[upload_num]) #save the new file as the same name so it is replaced

    ##### VERIFY COLUMNS #####
    for x in range(len(EXPECTED_UPLOAD_COLUMNS)):
        if data.columns[x] != EXPECTED_UPLOAD_COLUMNS[x]:
            print("Error with column " + str(x))
            print("Expected: " + EXPECTED_UPLOAD_COLUMNS[x])
            print("Upload: " + data.columns[x])
            print("Terminating program- Adjust upload file and run program again.")
            sys.exit()

    total_qty = data['Sort Position'].max() #gets the tota195441 DZS_Intl - Copyl amt of rows in Sort Position column
    print("Total Qty: ", end ="")
    print(total_qty)
    total_qty_array.append(total_qty)

    mid_rec = int((total_qty-1) / 2) #middle record in the data file, divides by 2, subtracts 1

    #### Job Number ####
    print("Job Number: ", end ="")
    print(job_num)

    #### Job File Name ####
    print("Job Name: ", end ="")
    print(job_info_array[upload_num])

    #### Job Permit Number ####
    print("Permit Number: ", end ="")
    print(permit_num)

    #### Job Mail Class ####
    print(mail_class)

    #### record names at end of data file ####
    lr_name = data['Full Name'].loc[data.index[total_qty-1]] #full name column, row (total_qty-1)
    fr_name = data['Full Name'].loc[data.index[0]] #full name column, row 1 full name
    
    print("Last record name: ", end ="")
    print(lr_name)

    #### (veri)fication record information (pandas) ####
    print(f"Verification Record: (Record #{mid_rec})")
    veri_rec_name = data['Full Name'].loc[data.index[mid_rec]] #full name column, row mid_rec
    print('--Full Name.......' + str(veri_rec_name))

    veri_rec_biz = data['Business'].loc[data.index[mid_rec]] #business column, row mid_rec
    print('--Business:.......' + str(veri_rec_biz))

    veri_rec_addr1 = data['Address Line 1'].loc[data.index[mid_rec]] #address line 1 column, row mid_rec
    print('--Address 1.......' + str(veri_rec_addr1))

    veri_rec_addr2 = data['Address Line 2'].loc[data.index[mid_rec]] #address line 2 column, row mid_rec
    print('--Address 2.......' + str(veri_rec_addr2))

    veri_rec_csz = data['City State Zip Code'].loc[data.index[mid_rec]] #city state zip column, row mid_rec
    print('--City ST, ZIP....' + str(veri_rec_csz))

    veri_rec_sort_num = data['Sort Position'].loc[data.index[mid_rec]] #sort position column, row mid_rec
    print('--Sort............' + str(veri_rec_sort_num))


    ############ Postage Costs ############
    if permit_num=='95':
        mrkdup_postage = (MARKUP_PERCENT * float(total_qty))# + postage
        new_postage = float(postage) + mrkdup_postage
        print(f"The new marked up postage is: ${new_postage}")

    elif permit_num=='462':
        mrkdup_postage = (MARKUP_PERCENT * float(total_qty))# + postage
        new_postage = float(postage) + mrkdup_postage
        print(f"The new marked up postage is: ${new_postage}")

    elif permit_num=='NA':
        new_postage = postage
        
    elif permit_num=='INTL':
        new_postage = postage
        
    elif permit_num=='DP':
        mrkdup_postage = (MARKUP_PERCENT * float(total_qty))# + postage
        new_postage = float(postage) + mrkdup_postage
        print(f"The new marked up postage is: ${new_postage}")
    else:
        new_postage = postage

    postage_array.append(new_postage)

    ############ Variables ############
    variables_array = []
    for x in data.columns:
        if x in EXPECTED_UPLOAD_COLUMNS:
            continue                    #The current column is a standard upload column, check next
        else:
            variables_array.append(x)   #The current column is not a standard upload column, and also not blank
    if len(variables_array) > 0:
        print("Variables detected: ", end = "")
        print(variables_array)

    ############ Checklist Creation (openpyxl)############
    chklst = load_workbook('Checklist-Template.xlsx')
    chklst_sheet = chklst.active

    chklst_sheet['C1'] = job_info_array[upload_num]             #Job Info Name + Number
    chklst_sheet['P1'] = total_qty                              #total amt of records 
    chklst_sheet['J15'] = permit_num                            #job permit number
    chklst_sheet['L15'] = mail_class                            #mailing class
    chklst_sheet['N22'] = lr_name                               #last record name
    chklst_sheet['J22'] = fr_name                               #first record name 
    chklst_sheet['C16'] = veri_rec_name                         #full name for verification record
    chklst_sheet['C18'] = veri_rec_biz                          #business for verification record
    chklst_sheet['C19'] = veri_rec_addr1                        #address line 1 for verification record
    chklst_sheet['C20'] = veri_rec_addr2                        #address line 2 for verification record
    chklst_sheet['C21'] = veri_rec_csz                          #city state zip code for verification record
    #chklst_sheet['D24'] = veri_rec_srt_tray_bun                #sort, tray bundle for verification record
    chklst_variable_cells = ['E29', 'E30', 'E31', 'E32', 'E33', 'E34', 'E35', 'E36', 'E37'] #list of all cells that CAN contain variable fields
    for x in range(len(variables_array)):
        if x > 9:
            break           #Will automatically stop on the 10th variable field, since there are only 9 possible variable fields available
        chklst_sheet[chklst_variable_cells[x]] = variables_array[x]


    if permit_num=='95':
        chklst_sheet['L27'] = total_qty #total amt of records
        
    elif permit_num=='462':
        chklst_sheet['L27'] = total_qty #total amt of records
        
    elif permit_num=='NA':  
        chklst_sheet['L28'] = total_qty #total amt of records
        
    elif permit_num=='INTL':  
        chklst_sheet['L29'] = total_qty #total amt of records
        
    elif permit_num=='DP':  
        chklst_sheet['L30'] = total_qty #total amt of records
        
        
    if permit_num=='95':
        chklst_sheet['O27'] = postage_array[upload_num] #new_postage cost into checklist
        
    elif permit_num=='462':
        chklst_sheet['O27'] = postage_array[upload_num] #new_postage cost into checklist
        
    elif permit_num=='NA':  
        chklst_sheet['O28'] = postage_array[upload_num] #postage cost into checklist
        
    elif permit_num=='INTL':  
        chklst_sheet['O29'] = postage_array[upload_num] #postage cost into checklist
        
    elif permit_num=='DP':  
        chklst_sheet['O30'] = postage_array[upload_num] #new_postage cost into checklist


    chklst_sheet.title = job_info_array[upload_num] #the title of the sheet would be job_info
    chklst.save(job_info_array[upload_num] + ' Checklist.xlsx') #save as new file name after job_info 

    chklst_name = job_info_array[upload_num] + ' Checklist.xlsx'

    ############ PRINT THE CHECKLIST FILE ############
    #should print checklist by accessing the current directory, finding the file/job name
    #then will print that file, hopefully whatever it is

    print("Printing checklist...")
    os.startfile(chklst_name, "print")

    ############ finishing touches for data (openpyxl) ############
    fi_data = load_workbook(filename_array[upload_num])#deletes the generated column from the pandas addin
    fi_sheet = fi_data.active
    fi_sheet.delete_cols(1)

    print("Adding 5 mail samples...")
    m1_row = total_qty + 2
    m2_row = total_qty + 3
    m3_row = total_qty + 4
    m4_row = total_qty + 5
    m5_row = total_qty + 6

    m1_name = fi_sheet.cell(row=m1_row, column=2) #row is the total_qty + 1, col 2 = full name
    m2_name = fi_sheet.cell(row=m2_row, column=2)
    m3_name = fi_sheet.cell(row=m3_row, column=2)
    m4_name = fi_sheet.cell(row=m4_row, column=2)
    m5_name = fi_sheet.cell(row=m5_row, column=2)

    m1_name.value = 'Mail Sample 1'
    m2_name.value = 'Mail Sample 2'
    m3_name.value = 'Mail Sample 3'
    m4_name.value = 'Mail Sample 4'
    m5_name.value = 'Mail Sample 5'

    m1_addr = fi_sheet.cell(row=m1_row, column=4)
    m2_addr = fi_sheet.cell(row=m2_row, column=4)
    m3_addr = fi_sheet.cell(row=m3_row, column=4)
    m4_addr = fi_sheet.cell(row=m4_row, column=4)
    m5_addr = fi_sheet.cell(row=m5_row, column=4)

    m1_addr.value = '4303 Normandy Ct'
    m2_addr.value = '4303 Normandy Ct'
    m3_addr.value = '4303 Normandy Ct'
    m4_addr.value = '4303 Normandy Ct'
    m5_addr.value = '4303 Normandy Ct'

    m1_csz = fi_sheet.cell(row=m1_row, column=6)
    m2_csz = fi_sheet.cell(row=m2_row, column=6)
    m3_csz = fi_sheet.cell(row=m3_row, column=6)
    m4_csz = fi_sheet.cell(row=m4_row, column=6)
    m5_csz = fi_sheet.cell(row=m5_row, column=6)

    m1_csz.value = 'Royal Oak MI 48073-2266'
    m2_csz.value = 'Royal Oak MI 48073-2266'
    m3_csz.value = 'Royal Oak MI 48073-2266'
    m4_csz.value = 'Royal Oak MI 48073-2266'
    m5_csz.value = 'Royal Oak MI 48073-2266'


    m1_imb = fi_sheet.cell(row=m1_row, column=7)
    m2_imb = fi_sheet.cell(row=m2_row, column=7)
    m3_imb = fi_sheet.cell(row=m3_row, column=7)
    m4_imb = fi_sheet.cell(row=m4_row, column=7)
    m5_imb = fi_sheet.cell(row=m5_row, column=7)

    m1_imb.value = 'FTTFTTFTDDADTTTFFTAADFTDTFATTFFDAAFADFFADFDTTATDTFADFDDDFTDAFTFFA'
    m2_imb.value = 'FTTFTTFTDDADTTTFFTAADFTDTFATTFFDAAFADFFADFDTTATDTFADFDDDFTDAFTFFA'
    m3_imb.value = 'FTTFTTFTDDADTTTFFTAADFTDTFATTFFDAAFADFFADFDTTATDTFADFDDDFTDAFTFFA'
    m4_imb.value = 'FTTFTTFTDDADTTTFFTAADFTDTFATTFFDAAFADFFADFDTTATDTFADFDDDFTDAFTFFA'
    m5_imb.value = 'FTTFTTFTDDADTTTFFTAADFTDTFATTFFDAAFADFFADFDTTATDTFADFDDDFTDAFTFFA'


    m1_numimb = fi_sheet.cell(row=m1_row, column=8)
    m2_numimb = fi_sheet.cell(row=m2_row, column=8)
    m3_numimb = fi_sheet.cell(row=m3_row, column=8)
    m4_numimb = fi_sheet.cell(row=m4_row, column=8)
    m5_numimb = fi_sheet.cell(row=m5_row, column=8)

    m1_numimb.value = '0027120244100016758199352413107'
    m2_numimb.value = '0027120244100016758199352413107'
    m3_numimb.value = '0027120244100016758199352413107'
    m4_numimb.value = '0027120244100016758199352413107'
    m5_numimb.value = '0027120244100016758199352413107'


    m1_oel = fi_sheet.cell(row=m1_row, column=9)
    m2_oel = fi_sheet.cell(row=m2_row, column=9)
    m3_oel = fi_sheet.cell(row=m3_row, column=9)
    m4_oel = fi_sheet.cell(row=m4_row, column=9)
    m5_oel = fi_sheet.cell(row=m5_row, column=9)


    m1_oel.value = '***************AUTO**MIXED AADC 480'
    m2_oel.value = '***************AUTO**MIXED AADC 480'
    m3_oel.value = '***************AUTO**MIXED AADC 480'
    m4_oel.value = '***************AUTO**MIXED AADC 480'
    m5_oel.value = '***************AUTO**MIXED AADC 480'

    m1_sp = fi_sheet.cell(row=m1_row, column=10)
    m2_sp = fi_sheet.cell(row=m2_row, column=10)
    m3_sp = fi_sheet.cell(row=m3_row, column=10)
    m4_sp = fi_sheet.cell(row=m4_row, column=10)
    m5_sp = fi_sheet.cell(row=m5_row, column=10)

    m1_sp.value = '0'
    m2_sp.value = '0'
    m3_sp.value = '0'
    m4_sp.value = '0'
    m5_sp.value = '0'

    m1_bun_num = fi_sheet.cell(row=m1_row, column=11)
    m2_bun_num = fi_sheet.cell(row=m2_row, column=11)
    m3_bun_num = fi_sheet.cell(row=m3_row, column=11)
    m4_bun_num = fi_sheet.cell(row=m4_row, column=11)
    m5_bun_num = fi_sheet.cell(row=m5_row, column=11)


    fi_data.save(filename_array[upload_num])

    more_to_upload = input("Do you have another file to upload? ")
    if more_to_upload == "Yes" or more_to_upload == "yes" or more_to_upload == "y" or more_to_upload == "Y":
        upload_num += 1
        continue
    else:
        break


############ Outlook Job Email ############
#Outlook Application Instance
olApp = win32.Dispatch('Outlook.Application')
olNS = olApp.GetNameSpace('MAPI')
os.getcwd()

# construct the email item object
mailItem = olApp.CreateItem(0)

mailItem.Subject = job_info_array[0] # subject is job_info
mailItem.BodyFormat = 1

attached_text = ""
if permit_num_array[0] == '95' or permit_num_array[0] == '462':
    attached_text = "Attached is the Undeliverables and the Postage Request."
else:
    attached_text = "Attached is the Undeliverables and the Presort Report."

email_body = f'''Hi,

{attached_text}

QTY {total_qty_array[0]} – Postage ${postage_array[0]} – Client {client}
'''

for x in range(upload_num > 0):          #first upload is count zero, so anything beyond zero indicates multiple uploads
    email_body = email_body + f'''QTY {total_qty_array[1]} – Postage ${postage_array[1]} – Client {client}'''

mailItem.Body = email_body


presort_report = job_num + ' PresortReports.pdf'
presort_path = pathlib.Path(presort_report)

undeliv = 'Undeliverables ' + job_info + '.xlsx'
undeliv_path = pathlib.Path(undeliv)

presort_path.absolute()
undeliv_path.absolute()

str(presort_path.absolute())
str(undeliv_path.absolute())

presort_absolute = str(presort_path.absolute())
undeliv_absolute = str(undeliv_path.absolute())

print(presort_absolute)
print(undeliv_absolute)


#ck_hb = 
#ag_dk =
#sl_dk =
#sp_pc = 
#sp_kf =
#ml_sk = 


#mailItem.To = ck_hb
#mailItem.Attachments.Add(presort_absolute) # adds the presort report as an attachment 
#mailItem.Attachments.Add(undeliv_absolute)
#mailItem.Attachments.Add(os.path.join(os.getcwd(), 'example.png')) # adds the Undeliverables as an attachment

mailItem.Display()

#### Printing Finished Checklist ####
#class openpyxl.worksheet.page.PageMargins(left=0.75, right=0.75, top=1, bottom=1, header=0.5, footer=0.5
#Information about page margins for view/print layouts. Standard values (in inches) left, right = 0.75 top, bottom = 1 header, footer = 0.5

#class openpyxl.worksheet.page.PrintOptions(horizontalCentered=None, verticalCentered=None, headings=None, gridLines=None, gridLinesSet=None
#Worksheet print options

#class openpyxl.worksheet.page.PrintPageSetup(worksheet=None, orientation=None, paperSize=None, scale=None, fitToHeight=None, fitToWidth=None, firstPageNumber=None, useFirstPageNumber=None, paperHeight=None, paperWidth=None, pageOrder=None, usePrinterDefaults=None, blackAndWhite=None, draft=None, cellComments=None, errors=None, horizontalDpi=None, verticalDpi=None, copies=None, id=None)
#Worksheet print page setup
